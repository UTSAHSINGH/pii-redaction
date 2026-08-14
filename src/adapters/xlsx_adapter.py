"""
xlsx_adapter.py
---------------
Microsoft Excel (.xlsx) spreadsheet adapter for PII Shield.
Preserves formulas, cell styles, dimensions, sheet names, and merged cells.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Dict, List, Set, Tuple
import uuid

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from adapters.base import DocumentAdapter, NormalizedDocument
from models import DocumentSegment, PIIMatch


class XlsxAdapter(DocumentAdapter):
    """Adapter for Excel spreadsheets (.xlsx)."""

    supported_extensions = [".xlsx", ".xlsm", ".xltx"]

    def extract(self, file_source: str | Path | bytes, filename: str = "spreadsheet.xlsx") -> NormalizedDocument:
        if isinstance(file_source, bytes):
            stream = io.BytesIO(file_source)
            raw_bytes = file_source
        else:
            raw_bytes = Path(file_source).read_bytes()
            stream = io.BytesIO(raw_bytes)

        wb = openpyxl.load_workbook(stream, data_only=False)
        segments: List[DocumentSegment] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws: Worksheet = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows()):
                for col_idx, cell in enumerate(row):
                    val = cell.value
                    if val is None:
                        continue
                    text_val = str(val)
                    # Ignore purely formula references unless string literals exist
                    if text_val.strip():
                        seg_id = f"sheet_{sheet_idx}_r{row_idx + 1}_c{col_idx + 1}"
                        segments.append(DocumentSegment(
                            segment_id=seg_id,
                            text=text_val,
                            metadata={
                                "sheet_name": sheet_name,
                                "sheet_idx": sheet_idx,
                                "row": row_idx + 1,
                                "col": col_idx + 1,
                                "coordinate": cell.coordinate,
                                "is_formula": isinstance(val, str) and val.startswith("="),
                            },
                            element_ref=cell,
                        ))

        return NormalizedDocument(
            document_id=str(uuid.uuid4()),
            filename=filename,
            file_type="xlsx",
            segments=segments,
            raw_bytes=raw_bytes,
            metadata={"sheets": len(wb.sheetnames), "sheet_names": wb.sheetnames},
            native_doc=wb,
        )

    def apply_redactions(
        self,
        doc: NormalizedDocument,
        approved_matches: List[PIIMatch],
    ) -> bytes:
        wb: openpyxl.Workbook = doc.native_doc
        matches_by_seg: Dict[str, List[PIIMatch]] = {}
        for m in approved_matches:
            matches_by_seg.setdefault(m.segment_id, []).append(m)

        for seg in doc.segments:
            seg_matches = matches_by_seg.get(seg.segment_id, [])
            if not seg_matches:
                continue

            cell = seg.element_ref
            if not cell:
                continue

            sorted_matches = sorted(seg_matches, key=lambda m: m.start, reverse=True)
            chars = list(seg.text)
            for m in sorted_matches:
                replacement = m.replacement or ""
                chars[m.start:m.end] = list(replacement)

            cell.value = "".join(chars)

        out_stream = io.BytesIO()
        wb.save(out_stream)
        return out_stream.getvalue()
