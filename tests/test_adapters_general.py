"""
test_adapters_general.py
------------------------
Unit and integration tests for multi-format document adapters:
DOCX, TXT, CSV, XLSX, and PDF.
"""

import io
import sys
from pathlib import Path
import pytest
from docx import Document
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from adapters.registry import get_adapter_for_file, get_supported_extensions
from models import PIIMatch, RedactionStrategy
from replacement_generator import get_replacement


def test_supported_extensions():
    exts = get_supported_extensions()
    for required in [".docx", ".txt", ".csv", ".xlsx", ".pdf"]:
        assert required in exts


def test_txt_adapter_extract_and_redact(tmp_path):
    txt_path = tmp_path / "sample.txt"
    txt_path.write_text("Hello John Doe, contact me at john@example.com.\nSecond line text.", encoding="utf-8")

    adapter = get_adapter_for_file(txt_path)
    doc = adapter.extract(txt_path, "sample.txt")

    assert len(doc.segments) == 2
    assert "John Doe" in doc.segments[0].text

    match = PIIMatch(
        match_id="m1",
        segment_id="line_0",
        entity_type="PERSON",
        start=6,
        end=14,
        text="John Doe",
        confidence=0.98,
        replacement="Alex Carter",
    )

    redacted_bytes = adapter.apply_redactions(doc, [match])
    redacted_text = redacted_bytes.decode("utf-8")
    assert "Hello Alex Carter, contact me at john@example.com." in redacted_text
    assert "Second line text." in redacted_text


def test_csv_adapter_extract_and_redact(tmp_path):
    csv_path = tmp_path / "sample.csv"
    csv_path.write_text("ID,Name,Email\n1,Sarah Connor,sarah@skynet.com\n", encoding="utf-8")

    adapter = get_adapter_for_file(csv_path)
    doc = adapter.extract(csv_path, "sample.csv")

    assert len(doc.segments) >= 3
    # Redact email
    m_email = PIIMatch(
        match_id="m_email",
        segment_id="csv_r1_c2",
        entity_type="EMAIL",
        start=0,
        end=len("sarah@skynet.com"),
        text="sarah@skynet.com",
        confidence=0.99,
        replacement="user@example.com",
    )

    redacted_bytes = adapter.apply_redactions(doc, [m_email])
    redacted_csv = redacted_bytes.decode("utf-8")
    assert "user@example.com" in redacted_csv
    assert "sarah@skynet.com" not in redacted_csv


def test_xlsx_adapter_extract_and_redact(tmp_path):
    xlsx_path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Employee Name"
    ws["B1"] = "Marcus Vance"
    wb.save(str(xlsx_path))

    adapter = get_adapter_for_file(xlsx_path)
    doc = adapter.extract(xlsx_path, "sample.xlsx")

    assert len(doc.segments) == 2
    match = PIIMatch(
        match_id="m1",
        segment_id="sheet_0_r1_c2",
        entity_type="PERSON",
        start=0,
        end=len("Marcus Vance"),
        text="Marcus Vance",
        confidence=0.95,
        replacement="Logan Sullivan",
    )

    redacted_bytes = adapter.apply_redactions(doc, [match])
    redacted_wb = openpyxl.load_workbook(io.BytesIO(redacted_bytes))
    assert redacted_wb.active["B1"].value == "Logan Sullivan"
